#!/usr/bin/env python3
"""
Mist Access Assurance — Deny Log Intelligence
Fetches RADIUS deny events, aggregates by client (MAC), and generates
a self-contained HTML dashboard + Excel workbook.

Usage:
    pip install requests openpyxl
    python3 deny_report.py
"""

import sys
import os
import re
import json
import getpass
import subprocess
from datetime import datetime, timezone, timedelta
from collections import defaultdict, Counter

try:
    import requests
except ImportError:
    requests = None

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

DENY_EVENT_TYPES = {"NAC_CLIENT_DENY", "NAC_SERVER_CERT_VALIDATION_FAILURE"}

CATEGORY_LABELS = {
    "cert": "Cert / TLS Issue",
    "cred": "Wrong Credentials",
    "mac":  "MAC Auth / Policy Failure",
}

CATEGORY_REMEDIATION = {
    "cert": (
        "Deploy updated client certificate via MDM. "
        "Cert chain must include the issuing CA trusted by the RADIUS server. "
        "Verify the client is configured to trust the Mist Authentication Service "
        "certificate (Organization > Access > Certificates)."
    ),
    "cred": (
        "Verify username/password are correct and the account is not locked or expired. "
        "Check that the IdP group membership allows network access. "
        "Re-enroll the device if credentials were recently rotated."
    ),
    "mac": (
        "Confirm the device MAC address is enrolled in the correct MAC Auth list. "
        "Review NAC policy rules — the device is hitting the implicit deny. "
        "Check if the SSID or port policy requires a matching rule for this device type."
    ),
}

# ─────────────────────────────────────────────────────────────────────────────
# Diagnosis engine — maps raw RADIUS error text → specific root cause + fix
# Each entry: (regex_pattern, reason_string, fix_string)
# Evaluated top-to-bottom; first match wins.
# ─────────────────────────────────────────────────────────────────────────────

DENY_DIAGNOSIS = [
    # ── Unsupported EAP methods ───────────────────────────────────────────────
    (
        r"eap.?peap|peap",
        "Device configured for EAP-PEAP — not supported by Mist Access Assurance",
        "Mist Access Assurance uses EAP-TLS (certificate-based) only. The device's "
        "wireless profile must be changed from PEAP to EAP-TLS. Push an updated Wi-Fi "
        "profile via MDM that selects EAP-TLS and references the device certificate.",
    ),
    (
        r"eap.?md5",
        "Device using EAP-MD5 — legacy method not supported by Mist Access Assurance",
        "EAP-MD5 is a deprecated, insecure method not supported by Mist AA. Update the "
        "wireless profile to EAP-TLS via MDM and ensure a device certificate is enrolled.",
    ),
    (
        r"eap.?fast",
        "Device using EAP-FAST — not supported by Mist Access Assurance",
        "EAP-FAST is not supported. Update the wireless profile to EAP-TLS via MDM.",
    ),
    (
        r"eap.?ttls|eap.?sim|eap.?aka|eap.?gtc|eap.?pwd",
        "Device using an unsupported EAP method",
        "Mist Access Assurance requires EAP-TLS. Update the wireless profile to EAP-TLS "
        "via MDM and ensure a device certificate is enrolled.",
    ),
    # ── Certificate / TLS errors ──────────────────────────────────────────────
    (
        r"(certificate verify failed|cert.*verify|verify.*cert).*(unknown ca|unable to get local issuer|no trusted|untrusted|ca.*not|issuer.*not)",
        "Client does not trust the RADIUS server certificate (unknown / untrusted CA)",
        "The device has not received the Mist org CA certificate. Export it from "
        "Organization > Access > Certificates and deploy as a Trusted Root via MDM. "
        "Jamf: Configuration Profiles > Certificate payload. "
        "Intune: Trusted Certificate profile.",
    ),
    (
        r"certificate.*expir|cert.*expir|expir.*certif",
        "Client certificate has expired",
        "The device certificate has passed its validity end date. Issue a new certificate "
        "from your PKI and push it via MDM. Review MDM certificate renewal policies to "
        "prevent this recurring.",
    ),
    (
        r"certificate.*not.*yet.*valid|not yet valid",
        "Client certificate is not yet valid — likely a clock skew issue",
        "The certificate's NotBefore date is in the future. Verify the device clock is "
        "synced (NTP). If the clock is correct, the cert was issued with a future start "
        "date — reissue it.",
    ),
    (
        r"certificate.*revoked|revoked.*cert|ocsp|crl.*fail",
        "Client certificate has been revoked",
        "The certificate is on a revocation list (CRL/OCSP). Issue a replacement cert and "
        "push via MDM. Investigate why it was revoked — device re-enrollment may be needed.",
    ),
    (
        r"no certificate|missing.*cert|cert.*missing|no client cert|empty.*cert",
        "No client certificate presented — EAP-TLS requires a device certificate",
        "The device attempted EAP-TLS but sent no certificate. Enroll a certificate via MDM "
        "(SCEP or PKCS#12 profile) and verify the wireless profile references it as the "
        "client authentication certificate.",
    ),
    (
        r"certificate.*chain|chain.*cert|intermediate|issuing ca",
        "Incomplete certificate chain — intermediate CA may be missing",
        "The client certificate chain is incomplete. The device or RADIUS server is missing "
        "an intermediate CA certificate. Ensure the full chain (leaf + intermediate + root) "
        "is included in the MDM certificate payload.",
    ),
    (
        r"tls.*alert|alert.*handshake|handshake.*fail|tls.*error|ssl.*error|tls.*fatal",
        "TLS handshake failed — certificate or protocol negotiation error",
        "A TLS-level error prevented authentication. Common causes: "
        "1) Device does not trust the server cert — deploy Mist CA via MDM. "
        "2) Cipher suite or TLS version mismatch — check device supplicant settings. "
        "3) Client cert is malformed or missing the clientAuth Extended Key Usage (EKU).",
    ),
    # ── Credential / LDAP / IdP errors ───────────────────────────────────────
    (
        r"invalid credentials|invalid password|wrong password|bind.*fail|ldap.*bind",
        "Username or password rejected by the directory (LDAP/AD)",
        "The IdP explicitly rejected the credentials. Verify the username and password are "
        "correct. Check if the password was recently changed without updating the device's "
        "saved network credentials. Look for account lockout if failures are repeated.",
    ),
    (
        r"unknown user|no such user|user.*not.*found|no.*user.*found|user.*unknown|no.*account",
        "User account not found in directory",
        "The username does not exist in the configured IdP/LDAP. "
        "Verify the username format (UPN vs. sAMAccountName). "
        "Check that the LDAP search base in Mist (Org > Access > Identity Providers) "
        "covers the user's OU. Confirm the account has not been deleted.",
    ),
    (
        r"account.*disabled|user.*disabled|account.*inactive|disabled.*account|user.*inactive",
        "User account is disabled in the directory",
        "The IdP reports this account as disabled. Re-enable it in Active Directory / "
        "Entra ID, or re-enroll the device under an active account.",
    ),
    (
        r"account.*locked|locked.*account|too many.*attempt|account.*blocked|intruder.*lockout",
        "User account is locked out",
        "Too many failed authentication attempts have locked the account. Unlock it in "
        "Active Directory / Entra ID. Investigate why the device was generating repeated "
        "failures — the saved Wi-Fi password profile is likely outdated.",
    ),
    (
        r"password.*expir|expir.*password|must.*change.*password|password.*change.*required",
        "User password has expired",
        "The user must reset their password. Once reset, update saved Wi-Fi credentials on "
        "the device. For domain-joined devices, ensure the device can reach a domain "
        "controller to pick up the new credentials.",
    ),
    (
        r"not.*member|group.*member|no.*group.*access|group.*not.*found|authorization.*failed|not.*authoriz",
        "User authenticated but not authorized — missing group membership",
        "The IdP confirmed the user's identity but they are not in a group that grants "
        "network access. Verify the NAC policy rule in Mist — check which IdP attribute "
        "(group / OU / role) is being matched. Add the user to the required group in "
        "Active Directory / Entra ID.",
    ),
    # ── MAC / Policy errors ───────────────────────────────────────────────────
    (
        r"no policy.*rules.*matched|implicit deny|no.*rule.*matched|policy.*not.*matched|no matching.*rule",
        "No NAC policy rule matched this device — hit implicit deny",
        "The device reached the end of all NAC rules without a match. "
        "1) Verify the correct SSID / port policy is applied. "
        "2) Review Organization > Access > NAC Rules — a rule must match on MAC label, "
        "certificate, or IdP attribute. "
        "3) For MAC auth devices, confirm the MAC is in the correct label/list.",
    ),
    (
        r"mac.*not.*found|mac.*not.*list|mac.*unknown|unknown.*mac|not.*mac.*auth|mac.*auth.*fail",
        "MAC address not enrolled in any MAC Authentication list",
        "This device's MAC is not in any MAC Auth client list in Mist. Add it under "
        "Organization > Access > Client Lists, or create a NAC rule that authenticates "
        "this device via certificate or IdP instead.",
    ),
    (
        r"vlan.*not.*found|vlan.*fail|no.*vlan|vlan.*assign|vlan.*not.*exist",
        "VLAN assignment failed — VLAN may not exist on the switch or AP",
        "Authentication succeeded but the assigned VLAN could not be applied. Verify the "
        "VLAN ID in the matching NAC rule exists on the switch / AP. Check VLAN definitions "
        "in the network configuration.",
    ),
    (
        r"rate.?limit|too many request|throttl",
        "Authentication attempts are being rate-limited",
        "The client is sending authentication requests faster than the server allows. "
        "This usually means a device in a retry loop. Check the supplicant configuration "
        "for aggressive retry timers. Resolve the underlying auth failure to stop the loop.",
    ),
]


def diagnose_text(primary_text, has_server_cert_fail=False):
    """
    Match the primary deny reason text against DENY_DIAGNOSIS patterns.
    Returns (reason, fix) strings, or (None, None) if no pattern matches.
    NAC_SERVER_CERT_VALIDATION_FAILURE events indicate the client rejected
    the server cert, regardless of any accompanying text.
    """
    if has_server_cert_fail:
        return (
            "Client rejected the RADIUS server certificate (server cert not trusted)",
            "The device's supplicant failed to validate the Mist RADIUS server certificate. "
            "Export the Mist org CA from Organization > Access > Certificates and deploy it "
            "to devices as a Trusted Root via MDM. Without this, EAP-TLS cannot complete the "
            "handshake even if the client certificate is valid.",
        )
    tl = (primary_text or "").lower()
    if not tl:
        return None, None
    for pattern, reason, fix in DENY_DIAGNOSIS:
        if re.search(pattern, tl):
            return reason, fix
    return None, None


REGIONS = [
    ("Global          (api.mist.com)",       "https://api.mist.com"),
    ("Europe          (api.eu.mist.com)",     "https://api.eu.mist.com"),
    ("APAC            (api.gc1.mist.com)",    "https://api.gc1.mist.com"),
    ("Global 03       (api.gc2.mist.com)",    "https://api.gc2.mist.com"),
    ("Global 04       (api.gc3.mist.com)",    "https://api.gc3.mist.com"),
    ("Global 05       (api.gc4.mist.com)",    "https://api.gc4.mist.com"),
]

BIZ_HOUR_START      = 7   # 7 am local
BIZ_HOUR_END        = 19  # 7 pm local
SILENCE_BIZ_HOURS   = 8   # 1 full business day without retry = silent


# ─────────────────────────────────────────────────────────────────────────────
# Classification
# ─────────────────────────────────────────────────────────────────────────────

def classify_event(event):
    """Map a raw NAC event to cert | cred | mac."""
    etype = event.get("type", "")
    text  = event.get("text", "").lower()
    auth  = event.get("auth_type", "")
    if etype == "NAC_SERVER_CERT_VALIDATION_FAILURE":
        return "cert"
    if "certificate" in text or "cert" in text or "tls" in text or "ca " in text:
        return "cert"
    if auth == "device-auth":
        return "mac"
    if "policy rules" in text or "implicit deny" in text or "no policy" in text:
        return "mac"
    return "cred"


# ─────────────────────────────────────────────────────────────────────────────
# Business-hour silence detection
# ─────────────────────────────────────────────────────────────────────────────

def business_hours_elapsed(from_ts, to_ts):
    """Count Mon–Fri 7am–7pm hours elapsed between two epoch timestamps."""
    from_dt = datetime.fromtimestamp(from_ts, tz=timezone.utc)
    to_dt   = datetime.fromtimestamp(to_ts,   tz=timezone.utc)
    if to_dt <= from_dt:
        return 0.0
    hours, current = 0.0, from_dt
    while current < to_dt:
        next_hour   = current + timedelta(hours=1)
        seg_end     = min(next_hour, to_dt)
        duration_h  = (seg_end - current).total_seconds() / 3600.0
        if current.weekday() < 5 and BIZ_HOUR_START <= current.hour < BIZ_HOUR_END:
            hours += duration_h
        current = next_hour
    return hours


# ─────────────────────────────────────────────────────────────────────────────
# (Asset CSV import is handled client-side in the HTML dashboard.
#  Drag-and-drop or file picker — no re-run needed.)
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# DEAD CODE MARKER — load_asset_csv kept here for reference only.
# The function below is never called; asset matching runs in the browser.
# ─────────────────────────────────────────────────────────────────────────────
def load_asset_csv(path):
    """
    UNUSED — asset matching is now handled in the HTML dashboard.

    Load a CSV of known managed device MAC addresses.
    Auto-detects the MAC address column by scanning header names.
    Returns:
        asset_map  — dict of {normalized_mac: {name, owner, dept, raw_row}}
        mac_col    — name of the column used for MAC matching
        row_count  — total rows parsed
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader  = csv.DictReader(f)
        headers = reader.fieldnames or []
        if not headers:
            return {}, None, 0

        mac_col   = _best_column(headers, _MAC_KEYWORDS)
        name_col  = _best_column(headers, _NAME_KEYWORDS)
        owner_col = _best_column(headers, _OWNER_KEYWORDS)
        dept_col  = _best_column(headers, _DEPT_KEYWORDS)

        if not mac_col:
            print(f"\n  Could not auto-detect MAC column.")
            print(f"  Columns found: {', '.join(headers)}")
            mac_col = input("  Enter the column name that contains MAC addresses: ").strip()
            if mac_col not in headers:
                print(f"  Column '{mac_col}' not found. Skipping asset matching.")
                return {}, None, 0

        asset_map = {}
        row_count = 0
        for row in reader:
            raw_mac = row.get(mac_col, "").strip()
            norm    = normalize_mac(raw_mac)
            if not norm or len(norm) != 12:
                continue
            row_count += 1
            asset_map[norm] = {
                "name":  row.get(name_col,  "").strip() if name_col  else "",
                "owner": row.get(owner_col, "").strip() if owner_col else "",
                "dept":  row.get(dept_col,  "").strip() if dept_col  else "",
            }

    return asset_map, mac_col, row_count


# ─────────────────────────────────────────────────────────────────────────────
# Aggregation
# ─────────────────────────────────────────────────────────────────────────────

def aggregate_events(events, site_map, lookback_days=7):
    """Collapse raw events to one record per client MAC."""
    now_ts          = datetime.now(tz=timezone.utc).timestamp()
    window_start_ts = now_ts - (lookback_days * 86400)

    clients     = defaultdict(lambda: {
        "mac": "", "site_id": "", "site": "", "ssid": "", "username": "",
        "auth_type": "", "category": None,
        "firstSeen": None, "lastSeen": None, "attempts": 0,
        "_days_active": set(), "_day_counts": defaultdict(int),
        "_text_counts": defaultdict(int),
        # last-known location fields (updated to most-recent deny event)
        "_last_ts": 0,
        "ap": "", "ap_mac": "", "port_id": "", "switch_mac": "",
        # diagnosis flag
        "_server_cert_fail": False,
    })
    permit_macs = set()

    for event in events:
        etype = event.get("type", "")
        ts    = float(event.get("timestamp", 0))
        if etype == "NAC_CLIENT_PERMIT":
            permit_macs.add(event.get("mac", ""))
            continue
        if etype not in DENY_EVENT_TYPES:
            continue
        if ts < window_start_ts:
            continue

        # Skip Marvis Mini synthetic test authentications — these are
        # infrastructure health checks on the management plane (port_type=vty),
        # not real client failures. They have no MAC address and always hit
        # the implicit deny by design. Including them pollutes the report.
        if event.get("port_type") == "vty":
            continue
        if not event.get("mac") and event.get("auth_type") == "device-auth":
            continue

        mac    = event.get("mac") or event.get("username") or "unknown"
        day_key = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d")
        c = clients[mac]
        c["mac"]       = mac
        c["site_id"]   = event.get("site_id", "")
        c["site"]      = site_map.get(event.get("site_id", ""), event.get("site_id", "unknown"))
        c["ssid"]      = event.get("ssid", "") or c["ssid"]
        c["username"]  = event.get("username", "") or c["username"]
        c["auth_type"] = event.get("auth_type", "") or c["auth_type"]
        c["attempts"] += 1
        c["_days_active"].add(day_key)
        c["_day_counts"][day_key] += 1
        if event.get("text"):
            c["_text_counts"][event["text"]] += 1
        if c["firstSeen"] is None or ts < c["firstSeen"]:
            c["firstSeen"] = ts
        if c["lastSeen"]  is None or ts > c["lastSeen"]:
            c["lastSeen"]  = ts

        # Keep last-known location from the most-recent deny event
        if ts > c["_last_ts"]:
            c["_last_ts"]   = ts
            c["ap"]         = event.get("ap", "") or event.get("ap_name", "") or ""
            c["ap_mac"]     = event.get("ap_mac", "") or ""
            c["port_id"]    = event.get("port_id", "") or event.get("port", "") or ""
            c["switch_mac"] = event.get("switch_mac", "") or ""

        if etype == "NAC_SERVER_CERT_VALIDATION_FAILURE":
            c["_server_cert_fail"] = True

        cat = classify_event(event)
        priority = {"cert": 3, "cred": 2, "mac": 1}
        if c["category"] is None or priority.get(cat, 0) > priority.get(c["category"], 0):
            c["category"] = cat

    # Build day labels (oldest → newest)
    day_labels = []
    for i in range(lookback_days - 1, -1, -1):
        d = datetime.now(tz=timezone.utc) - timedelta(days=i)
        day_labels.append(d.strftime("%Y-%m-%d"))

    result = []
    for mac, c in clients.items():
        days_failing = len(c["_days_active"])
        text_counts  = dict(c["_text_counts"])
        primary_text = max(text_counts, key=text_counts.get) if text_counts else ""
        all_texts    = sorted(text_counts.items(), key=lambda x: -x[1])
        activity     = {day: c["_day_counts"].get(day, 0) for day in day_labels}
        biz_hours    = business_hours_elapsed(c["lastSeen"] or now_ts, now_ts)

        if mac in permit_macs:
            status = "resolved"
        else:
            status = "silent" if biz_hours >= SILENCE_BIZ_HOURS else "failing"

        cat = c["category"] or "mac"
        diagnosis, specific_fix = diagnose_text(primary_text, c["_server_cert_fail"])

        result.append({
            "mac":          mac,
            "site_id":      c["site_id"],
            "site":         c["site"],
            "ssid":         c["ssid"],
            "username":     c["username"],
            "auth_type":    c["auth_type"],
            "category":     cat,
            "categoryLabel": CATEGORY_LABELS.get(cat, ""),
            "remediation":  CATEGORY_REMEDIATION.get(cat, ""),
            "firstSeen":    c["firstSeen"],
            "lastSeen":     c["lastSeen"],
            "daysFailing":  days_failing,
            "attempts":     c["attempts"],
            "status":       status,
            "blastScore":   "low",
            "blastCount":   1,
            "activity":     activity,
            "primaryText":  primary_text,
            "allTexts":     [{"text": t, "count": n} for t, n in all_texts],
            "diagnosis":    diagnosis or "",
            "specificFix":  specific_fix or CATEGORY_REMEDIATION.get(cat, ""),
            "assetStatus":  "unknown",
            "assetName":    "",
            "assetOwner":   "",
            "assetDept":    "",
            "ap":           c["ap"],
            "apMac":        c["ap_mac"],
            "portId":       c["port_id"],
            "switchMac":    c["switch_mac"],
        })

    # Blast radius — how many clients share the same primary deny reason?
    reason_counts = Counter(c["primaryText"] for c in result if c["primaryText"])
    for c in result:
        shared = reason_counts.get(c["primaryText"], 1)
        c["blastCount"] = shared
        c["blastScore"] = "high" if shared >= 5 else ("med" if shared >= 3 else "low")

    result.sort(key=lambda x: (x["blastCount"], x["daysFailing"]), reverse=True)

    # Deny reasons breakdown (for the Deny Reasons tab)
    reason_map = {}
    for c in result:
        for item in c["allTexts"]:
            t = item["text"]
            if t not in reason_map:
                reason_map[t] = {"text": t, "clients": [], "totalEvents": 0}
            reason_map[t]["clients"].append({
                "mac": c["mac"], "username": c["username"],
                "site": c["site"], "count": item["count"],
                "category": c["category"],
            })
            reason_map[t]["totalEvents"] += item["count"]

    deny_reasons = sorted(
        [{"text": v["text"], "clients": v["clients"],
          "clientCount": len(v["clients"]), "totalEvents": v["totalEvents"]}
         for v in reason_map.values()],
        key=lambda x: (-x["clientCount"], -x["totalEvents"])
    )

    return result, day_labels, deny_reasons


# ─────────────────────────────────────────────────────────────────────────────
# Mist API
# ─────────────────────────────────────────────────────────────────────────────

def _headers(token):
    return {"Authorization": f"Token {token}", "Content-Type": "application/json"}


def fetch_org_info(token, base_url):
    """Call /api/v1/self, list all orgs, and let the user pick one."""
    resp = requests.get(f"{base_url}/api/v1/self", headers=_headers(token), timeout=30)
    resp.raise_for_status()
    data  = resp.json()
    privs = data.get("privileges", [])
    orgs  = [p for p in privs if p.get("scope") == "org"]

    if not orgs:
        print("\n  ERROR: No org-level access found for this token.")
        print("  An org-scoped admin token is required.")
        sys.exit(1)

    # Single org — skip the prompt
    if len(orgs) == 1:
        org = orgs[0]
        return {"org_id": org["org_id"], "org_name": org.get("name", org["org_id"])}

    # Multiple orgs — show a numbered list
    print("\nOrganizations available with this token:")
    for i, o in enumerate(orgs, 1):
        print(f"  {i}. {o.get('name', o['org_id'])}  ({o['org_id']})")

    while True:
        choice = input(f"Select org [1–{len(orgs)}]: ").strip()
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(orgs):
                org = orgs[idx]
                return {"org_id": org["org_id"], "org_name": org.get("name", org["org_id"])}
        except ValueError:
            pass
        print(f"  Please enter a number between 1 and {len(orgs)}.")


def fetch_sites(token, org_id, base_url):
    """Return dict of {site_id: site_name}."""
    resp = requests.get(
        f"{base_url}/api/v1/orgs/{org_id}/sites",
        headers=_headers(token), timeout=30
    )
    resp.raise_for_status()
    sites = resp.json()
    return {s["id"]: s["name"] for s in sites}


def fetch_events(token, org_id, lookback_days, base_url):
    """
    Paginate through NAC client events for the lookback window.
    Fetches deny types and permits separately so the API only returns
    relevant events — avoids downloading the full NAC event stream.
    """
    start_ts = int(datetime.now(tz=timezone.utc).timestamp()) - (lookback_days * 86400)
    url_base = f"{base_url}/api/v1/orgs/{org_id}/nac_clients/events/search"

    # Fetch deny types + permit in two focused passes rather than one
    # unfiltered pass. This can reduce total pages dramatically in large orgs.
    fetch_types = [
        ("NAC_CLIENT_DENY",                    "deny events"),
        ("NAC_SERVER_CERT_VALIDATION_FAILURE",  "cert failures"),
        ("NAC_CLIENT_PERMIT",                   "permit events"),
    ]

    all_events = []

    for event_type, label in fetch_types:
        type_events = []
        url    = url_base
        params = {"limit": 1000, "start": start_ts, "type": event_type}
        page   = 0

        while True:
            page += 1
            resp = requests.get(url, headers=_headers(token), params=params, timeout=60)
            resp.raise_for_status()
            data    = resp.json()
            results = data.get("results", [])
            type_events.extend(results)

            total = data.get("total") or data.get("estimated_total") or len(type_events)
            pct   = min(100, round(len(type_events) / max(total, 1) * 100))
            print(f"  {label}: page {page} — {len(type_events):,} / ~{total:,} ({pct}%)", end="\r")

            if not data.get("next"):
                break
            nxt    = data["next"]
            params = {}
            url    = f"{base_url}{nxt}" if nxt.startswith("/") else nxt

        print(f"  {label}: {len(type_events):,} fetched.                          ")
        all_events.extend(type_events)

    print(f"  Total events fetched: {len(all_events):,}           ")
    return all_events


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(report, path):
    wb = openpyxl.Workbook()

    # ── Shared styles ────────────────────────────────────────────────────────
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL   = PatternFill("solid", fgColor="1A1D27")
    WRAP       = Alignment(wrap_text=True, vertical="top")
    TOP        = Alignment(vertical="top")
    THIN       = Side(style="thin", color="C0C0C0")
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    CAT_COLORS = {"cert": "F97316", "cred": "A855F7", "mac": "06B6D4"}
    STS_COLORS = {"failing": "EF4444", "silent": "6366F1", "resolved": "22C55E"}
    BLT_COLORS = {"high": "EF4444", "med": "F59E0B", "low": "22C55E"}

    def style_header_row(ws, cols):
        for col, (header, width) in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font       = HDR_FONT
            cell.fill       = HDR_FILL
            cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border     = BORDER
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def color_cell(cell, hex_color, text_color="000000"):
        cell.fill      = PatternFill("solid", fgColor=hex_color)
        cell.font      = Font(color=text_color, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="top")

    def fmt_ts(ts):
        if not ts:
            return ""
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")

    # ── Sheet 1: Client Summary ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Client Summary"
    cols = [
        ("MAC Address", 18), ("Username", 22), ("Site", 20), ("SSID", 16),
        ("Auth Type", 14), ("Category", 22), ("Status", 12),
        ("Blast Radius", 14), ("Clients w/ Same Reason", 22),
        ("Total Attempts", 16), ("Days Failing", 14),
        ("First Seen", 18), ("Last Seen", 18),
        ("Primary Deny Reason", 60),
    ]
    style_header_row(ws, cols)

    for r, c in enumerate(report["clients"], 2):
        ws.cell(r, 1,  c["mac"]).font         = Font(name="Courier New", size=10)
        ws.cell(r, 2,  c.get("username", ""))
        ws.cell(r, 3,  c.get("site", ""))
        ws.cell(r, 4,  c.get("ssid", ""))
        ws.cell(r, 5,  c.get("auth_type", ""))
        cat_cell = ws.cell(r, 6,  c["categoryLabel"])
        color_cell(cat_cell, CAT_COLORS.get(c["category"], "888888"), "FFFFFF")
        sts_cell = ws.cell(r, 7,  c["status"])
        color_cell(sts_cell, STS_COLORS.get(c["status"], "888888"), "FFFFFF")
        blt_cell = ws.cell(r, 8,  c["blastScore"].upper())
        color_cell(blt_cell, BLT_COLORS.get(c["blastScore"], "888888"), "FFFFFF")
        ws.cell(r, 9,  c["blastCount"])
        ws.cell(r, 10, c["attempts"])
        ws.cell(r, 11, c["daysFailing"])
        ws.cell(r, 12, fmt_ts(c["firstSeen"]))
        ws.cell(r, 13, fmt_ts(c["lastSeen"]))
        reason_cell = ws.cell(r, 14, c.get("primaryText", ""))
        reason_cell.alignment = WRAP
        ws.row_dimensions[r].height = 40

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # ── Sheet 2: Deny Reasons ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Deny Reasons")
    cols2 = [
        ("Deny Reason Text", 80), ("Client Count", 14), ("Total Events", 14),
        ("Category", 22), ("Affected MACs", 50),
    ]
    style_header_row(ws2, cols2)
    for r, dr in enumerate(report["denyReasons"], 2):
        ws2.cell(r, 1, dr["text"]).alignment = WRAP
        ws2.cell(r, 2, dr["clientCount"])
        ws2.cell(r, 3, dr["totalEvents"])
        cats = list({c["category"] for c in dr["clients"]})
        cat  = cats[0] if len(cats) == 1 else "mixed"
        cat_cell2 = ws2.cell(r, 4, CATEGORY_LABELS.get(cat, cat))
        color_cell(cat_cell2, CAT_COLORS.get(cat, "888888"), "FFFFFF")
        macs = ", ".join(c.get("username") or c["mac"] for c in dr["clients"])
        ws2.cell(r, 5, macs).alignment = WRAP
        ws2.row_dimensions[r].height   = 40
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}1"

    # ── Sheet 3: Daily Timeline ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Daily Timeline")
    day_labels = report["dayLabels"]
    hdr_cols   = [("MAC / User", 22), ("Site", 18)] + [(d, 11) for d in day_labels] + [("Total", 9)]
    style_header_row(ws3, hdr_cols)
    for r, c in enumerate(report["clients"], 2):
        lbl = c.get("username") or c["mac"]
        ws3.cell(r, 1, lbl).font = Font(name="Courier New", size=10)
        ws3.cell(r, 2, c.get("site", ""))
        total = 0
        for col_i, day in enumerate(day_labels, 3):
            n = c.get("activity", {}).get(day, 0)
            cell = ws3.cell(r, col_i, n if n else "")
            if n:
                cell.fill      = PatternFill("solid", fgColor=CAT_COLORS.get(c["category"], "CCCCCC"))
                cell.font      = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            total += n
        ws3.cell(r, len(day_labels) + 3, total)

    # ── Sheet 4: Remediation Guide ───────────────────────────────────────────
    ws4 = wb.create_sheet("Remediation Guide")
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 30
    ws4.column_dimensions["C"].width = 80
    headers = ["Category", "Label", "Recommended Action"]
    for col, h in enumerate(headers, 1):
        cell = ws4.cell(1, col, h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r, (cat, label) in enumerate([
        ("cert", CATEGORY_LABELS["cert"]),
        ("cred", CATEGORY_LABELS["cred"]),
        ("mac",  CATEGORY_LABELS["mac"]),
    ], 2):
        ws4.cell(r, 1, cat)
        lc = ws4.cell(r, 2, label)
        color_cell(lc, CAT_COLORS[cat], "FFFFFF")
        rc = ws4.cell(r, 3, CATEGORY_REMEDIATION[cat])
        rc.alignment = WRAP
        ws4.row_dimensions[r].height = 60

    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# HTML Template  (static — data embedded at generation time)
# ─────────────────────────────────────────────────────────────────────────────

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Deny Log Intelligence</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root {
    --bg: #0f1117; --surface: #1a1d27; --surface2: #22263a;
    --border: #2e3350; --text: #e2e8f0; --text-muted: #8892a4;
    --accent: #4f8ef7; --cert: #f97316; --cred: #a855f7; --mac: #06b6d4;
    --high: #ef4444; --med: #f59e0b; --low: #22c55e;
    --silent: #6366f1; --resolved: #22c55e; --failing: #ef4444;
  }
  body { background: var(--bg); color: var(--text); font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; font-size: 14px; min-height: 100vh; }
  .header { background: var(--surface); border-bottom: 1px solid var(--border); padding: 14px 24px; display: flex; align-items: center; justify-content: space-between; }
  .header-title { font-size: 18px; font-weight: 700; }
  .header-meta { color: var(--text-muted); font-size: 12px; }
  .header-right { display: flex; align-items: center; gap: 16px; font-size: 12px; color: var(--text-muted); }
  .print-btn { padding: 5px 12px; border: 1px solid var(--border); border-radius: 6px; background: var(--surface2); color: var(--text); cursor: pointer; font-size: 12px; }
  .print-btn:hover { border-color: var(--accent); color: var(--accent); }
  .import-btn { padding: 5px 12px; border: 1px solid #16a34a; border-radius: 6px; background: rgba(34,197,94,0.1); color: #4ade80; cursor: pointer; font-size: 12px; }
  .import-btn:hover { background: rgba(34,197,94,0.2); }
  .asset-bar { background: rgba(34,197,94,0.08); border-bottom: 1px solid rgba(34,197,94,0.25); padding: 8px 24px; display:flex; align-items:center; gap:12px; font-size:12px; }
  .asset-bar-icon { font-size:16px; }
  .asset-bar-text { color: #4ade80; flex:1; }
  .asset-bar-clear { color: var(--text-muted); cursor:pointer; text-decoration:underline; }
  .asset-bar-clear:hover { color: var(--text); }
  .toast { position:fixed; bottom:24px; right:24px; background:#1e293b; border:1px solid var(--border); border-radius:8px; padding:12px 18px; font-size:13px; z-index:9999; opacity:0; transition:opacity 0.2s; pointer-events:none; max-width:380px; }
  .toast.show { opacity:1; }
  .toast.toast-ok   { border-color:#16a34a; color:#4ade80; }
  .toast.toast-err  { border-color:#dc2626; color:#f87171; }
  .toast.toast-warn { border-color:#d97706; color:#fbbf24; }
  .drop-zone-active { outline: 2px dashed #4ade80 !important; }
  .main { padding: 24px; max-width: 1400px; margin: 0 auto; }
  .cards-row { display: grid; gap: 16px; margin-bottom: 24px; }
  .cards-row.metric-cards { grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); }
  .cards-row.category-cards { grid-template-columns: repeat(3, 1fr); }
  .card { background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 18px 20px; }
  .card-label { font-size: 11px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 8px; }
  .card-value { font-size: 28px; font-weight: 700; line-height: 1; }
  .card-sub { font-size: 12px; color: var(--text-muted); margin-top: 6px; }
  .card.clickable { cursor: pointer; transition: border-color 0.15s; }
  .card.clickable:hover { border-color: var(--accent); }
  .card.active-filter { border-color: var(--accent); background: rgba(79,142,247,0.08); }
  .cat-icon { font-size: 22px; margin-bottom: 8px; }
  .cat-count { font-size: 26px; font-weight: 700; }
  .cat-label { font-size: 12px; color: var(--text-muted); margin-top: 4px; }
  .cert-color { color: var(--cert); } .cred-color { color: var(--cred); } .mac-color { color: var(--mac); }
  .section { background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 20px; margin-bottom: 24px; }
  .section-title { font-size: 14px; font-weight: 600; margin-bottom: 16px; }
  .timeline-header { display: flex; gap: 0; margin-bottom: 6px; padding-left: 200px; }
  .timeline-day-label { flex: 1; text-align: center; font-size: 11px; color: var(--text-muted); min-width: 40px; }
  .timeline-row { display: flex; align-items: center; margin-bottom: 4px; }
  .timeline-client-label { width: 200px; font-size: 12px; color: var(--text-muted); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; padding-right: 12px; flex-shrink: 0; }
  .timeline-cells { display: flex; gap: 3px; flex: 1; }
  .timeline-cell { flex: 1; min-width: 40px; height: 20px; border-radius: 3px; background: var(--surface2); }
  .timeline-cell.cert { background: var(--cert); }
  .timeline-cell.cred { background: var(--cred); }
  .timeline-cell.mac  { background: var(--mac); }
  .reason-row { display: flex; align-items: flex-start; gap: 12px; padding: 12px 0; border-bottom: 1px solid var(--border); cursor: pointer; }
  .reason-row:last-child { border-bottom: none; }
  .reason-row:hover { background: var(--surface2); margin: 0 -20px; padding: 12px 20px; border-radius: 6px; }
  .reason-row.active-reason { background: rgba(79,142,247,0.08); margin: 0 -20px; padding: 12px 20px; border-radius: 6px; border-left: 3px solid var(--accent); }
  .reason-bar-wrap { width: 130px; flex-shrink: 0; display: flex; align-items: center; gap: 8px; }
  .reason-bar { height: 8px; border-radius: 4px; min-width: 4px; }
  .reason-client-count { font-size: 12px; color: var(--text-muted); white-space: nowrap; }
  .reason-text { font-size: 13px; flex: 1; line-height: 1.5; }
  .reason-macs { font-size: 11px; color: var(--text-muted); margin-top: 3px; font-family: monospace; }
  .alert-card { background: rgba(239,68,68,0.08); border: 1px solid rgba(239,68,68,0.3); border-radius: 8px; padding: 14px 16px; margin-bottom: 12px; }
  .alert-card.silent { background: rgba(99,102,241,0.08); border-color: rgba(99,102,241,0.3); }
  .alert-title { font-weight: 600; font-size: 13px; margin-bottom: 4px; }
  .alert-body { font-size: 12px; color: var(--text-muted); }
  .table-wrapper { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { text-align: left; padding: 10px 12px; border-bottom: 2px solid var(--border); color: var(--text-muted); font-weight: 600; font-size: 11px; text-transform: uppercase; letter-spacing: 0.04em; cursor: pointer; white-space: nowrap; user-select: none; }
  th:hover { color: var(--text); }
  td { padding: 10px 12px; border-bottom: 1px solid var(--border); vertical-align: middle; }
  tr:hover td { background: var(--surface2); }
  .mac-cell { font-family: monospace; font-size: 12px; }
  .badge { display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 11px; font-weight: 600; }
  .badge-cert     { background: rgba(249,115,22,0.15);  color: var(--cert); }
  .badge-cred     { background: rgba(168,85,247,0.15);  color: var(--cred); }
  .badge-mac      { background: rgba(6,182,212,0.15);   color: var(--mac);  }
  .badge-high     { background: rgba(239,68,68,0.15);   color: var(--high); }
  .badge-med      { background: rgba(245,158,11,0.15);  color: var(--med);  }
  .badge-low      { background: rgba(34,197,94,0.15);   color: var(--low);  }
  .badge-failing  { background: rgba(239,68,68,0.15);   color: var(--high); }
  .badge-silent   { background: rgba(99,102,241,0.15);  color: var(--silent); }
  .badge-resolved   { background: rgba(34,197,94,0.15);   color: var(--resolved); }
  .badge-managed    { background: rgba(34,197,94,0.15);   color: #16a34a; }
  .badge-unmanaged  { background: rgba(156,163,175,0.15); color: #6b7280; }
  .badge-unknown    { background: rgba(156,163,175,0.10); color: #9ca3af; }
  .tabs { display: flex; gap: 4px; border-bottom: 1px solid var(--border); margin-bottom: 20px; }
  .tab { padding: 8px 16px; cursor: pointer; border-radius: 6px 6px 0 0; font-size: 13px; color: var(--text-muted); border: 1px solid transparent; border-bottom: none; }
  .tab.active { background: var(--surface); border-color: var(--border); color: var(--text); margin-bottom: -1px; }
  .tab-content { display: none; }
  .tab-content.active { display: block; }
  .filter-bar { display: flex; gap: 10px; align-items: center; margin-bottom: 16px; flex-wrap: wrap; }
  .filter-bar input  { background: var(--surface2); border: 1px solid var(--border); border-radius: 6px; padding: 6px 12px; color: var(--text); font-size: 13px; width: 240px; outline: none; }
  .filter-bar input:focus { border-color: var(--accent); }
  .filter-bar select { background: var(--surface2); border: 1px solid var(--border); border-radius: 6px; padding: 6px 10px; color: var(--text); font-size: 13px; outline: none; }
  .clear-btn { font-size: 12px; color: var(--text-muted); cursor: pointer; }
  .clear-btn:hover { color: var(--text); }
  .result-count { font-size: 12px; color: var(--text-muted); margin-left: auto; }
  .site-selector { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 20px; }
  .site-btn { padding: 6px 14px; border: 1px solid var(--border); border-radius: 20px; cursor: pointer; background: var(--surface2); color: var(--text); font-size: 12px; }
  .site-btn.active { border-color: var(--accent); color: var(--accent); }
  .email-box { background: var(--surface2); border: 1px solid var(--border); border-radius: 8px; padding: 16px; }
  .email-field { margin-bottom: 12px; }
  .email-field label { display: block; font-size: 11px; color: var(--text-muted); margin-bottom: 4px; text-transform: uppercase; }
  .email-field input, .email-field textarea { width: 100%; background: var(--bg); border: 1px solid var(--border); border-radius: 6px; padding: 8px 12px; color: var(--text); font-size: 13px; font-family: inherit; outline: none; }
  .email-field textarea { min-height: 300px; resize: vertical; line-height: 1.6; }
  .btn-row { display: flex; gap: 8px; margin-top: 12px; }
  .btn { padding: 8px 16px; border-radius: 6px; font-size: 13px; font-weight: 600; cursor: pointer; border: none; }
  .btn-primary { background: var(--accent); color: #fff; }
  .btn-secondary { background: var(--surface2); color: var(--text); border: 1px solid var(--border); }
  .btn:hover { opacity: 0.85; }
  .help-section { margin-bottom: 28px; }
  .help-section h3 { font-size: 15px; font-weight: 600; margin-bottom: 10px; color: var(--text); }
  .help-section p, .help-section li { font-size: 13px; color: var(--text-muted); line-height: 1.8; }
  .help-section ul, .help-section ol { padding-left: 20px; }
  .help-section li { margin-bottom: 4px; }
  .help-table { width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 8px; }
  .help-table th { text-align: left; padding: 8px 12px; border-bottom: 2px solid var(--border); color: var(--text-muted); font-size: 11px; text-transform: uppercase; }
  .help-table td { padding: 10px 12px; border-bottom: 1px solid var(--border); vertical-align: top; line-height: 1.6; }
  .help-table td:first-child { font-weight: 600; white-space: nowrap; color: var(--text); }
  .help-table td:last-child { color: var(--text-muted); }
  @media (max-width: 1000px) { .cards-row.metric-cards { grid-template-columns: repeat(3, 1fr); } }
  @media (max-width: 700px)  { .cards-row.metric-cards { grid-template-columns: repeat(2, 1fr); } .cards-row.category-cards { grid-template-columns: 1fr; } }
  @media print { .print-btn, .filter-bar, .tabs { display: none !important; } .tab-content { display: block !important; } }
</style>
</head>
<body>

<div class="header">
  <div>
    <div class="header-title">Deny Log Intelligence — <span id="hdr-org"></span></div>
    <div class="header-meta" id="hdr-meta"></div>
  </div>
  <div class="header-right">
    <span>Mist Access Assurance</span>
    <button class="import-btn" onclick="document.getElementById('csv-file-input').click()">📋 Import Asset List</button>
    <button class="print-btn" onclick="window.print()">🖨 Print / Save PDF</button>
  </div>
</div>

<input type="file" id="csv-file-input" accept=".csv,.txt" style="display:none" onchange="handleCSVFile(this.files[0])">

<div id="asset-bar" class="asset-bar" style="display:none">
  <span class="asset-bar-icon">✅</span>
  <span class="asset-bar-text" id="asset-bar-text"></span>
  <span class="asset-bar-clear" onclick="clearAssets()">✕ Remove</span>
</div>

<div id="toast" class="toast"></div>

<div class="main">
  <div class="cards-row metric-cards">
    <div class="card"><div class="card-label">Total Deny Events</div><div class="card-value" id="s-events">—</div><div class="card-sub" id="s-events-sub"></div></div>
    <div class="card"><div class="card-label">Unique Clients</div><div class="card-value" id="s-clients">—</div><div class="card-sub">Deduplicated by MAC</div></div>
    <div class="card"><div class="card-label">Sites Affected</div><div class="card-value" id="s-sites">—</div><div class="card-sub" id="s-sites-names"></div></div>
    <div class="card"><div class="card-label">Cert Failures</div><div class="card-value cert-color" id="s-cert">—</div><div class="card-sub" id="s-cert-pct"></div></div>
    <div class="card"><div class="card-label">Silent Failures <span id="silent-help" style="cursor:help;color:var(--text-muted)" title="">ⓘ</span></div><div class="card-value" style="color:var(--silent)" id="s-silent">—</div><div class="card-sub">No retry in 8+ business hours</div></div>
    <div class="card" id="card-managed" style="display:none"><div class="card-label">Managed Assets Failing</div><div class="card-value" style="color:#dc2626" id="s-managed">—</div><div class="card-sub" id="s-managed-sub"></div></div>
  </div>

  <div class="cards-row category-cards">
    <div class="card clickable" onclick="toggleCat('cert')" id="cat-cert"><div class="cat-icon">🔐</div><div class="cat-count cert-color" id="cc-cert">—</div><div class="cat-label">Cert / TLS Issues</div></div>
    <div class="card clickable" onclick="toggleCat('cred')" id="cat-cred"><div class="cat-icon">🔑</div><div class="cat-count cred-color" id="cc-cred">—</div><div class="cat-label">Wrong Credentials</div></div>
    <div class="card clickable" onclick="toggleCat('mac')"  id="cat-mac" ><div class="cat-icon">📱</div><div class="cat-count mac-color"  id="cc-mac" >—</div><div class="cat-label">MAC Auth / Policy Failure</div></div>
  </div>

  <div id="alerts"></div>

  <div class="tabs">
    <div class="tab active"  onclick="switchTab('dashboard')">Dashboard</div>
    <div class="tab"         onclick="switchTab('reasons')">Deny Reasons</div>
    <div class="tab"         onclick="switchTab('timeline')">7-Day Timeline</div>
    <div class="tab"         onclick="switchTab('notifications')">Notification Center</div>
    <div class="tab"         onclick="switchTab('help')">? Help</div>
  </div>

  <!-- Dashboard tab -->
  <div class="tab-content active" id="tab-dashboard">
    <div class="section">
      <div class="section-title">Client Deny Summary</div>
      <div class="filter-bar">
        <input type="text" id="search" placeholder="Search MAC, username, site, reason..." oninput="renderTable()">
        <select id="f-cat"    onchange="renderTable()"><option value="">All Categories</option><option value="cert">Cert / TLS</option><option value="cred">Credentials</option><option value="mac">MAC / Policy</option></select>
        <select id="f-site"   onchange="renderTable()"><option value="">All Sites</option></select>
        <select id="f-status" onchange="renderTable()"><option value="">All Statuses</option><option value="failing">Failing</option><option value="silent">Silent</option><option value="resolved">Resolved</option></select>
        <select id="f-reason" onchange="renderTable()"><option value="">All Deny Reasons</option></select>
        <select id="f-asset" onchange="renderTable()" style="display:none"><option value="">All Device Types</option><option value="managed">Managed</option><option value="unmanaged">Unmanaged</option></select>
        <span class="clear-btn" onclick="clearFilters()">Clear filters</span>
        <span class="result-count" id="result-count"></span>
      </div>
      <div class="table-wrapper">
        <table>
          <thead><tr>
            <th onclick="sortBy('mac')">MAC / User</th>
            <th onclick="sortBy('site')">Site</th>
            <th onclick="sortBy('category')">Category</th>
            <th onclick="sortBy('primaryText')" style="min-width:220px">Deny Reason</th>
            <th onclick="sortBy('firstSeen')">First Seen</th>
            <th onclick="sortBy('lastSeen')">Last Seen</th>
            <th onclick="sortBy('daysFailing')">Days</th>
            <th onclick="sortBy('attempts')">Attempts</th>
            <th onclick="sortBy('blastScore')" title="Blast radius = how many clients share the same deny reason">Blast ⓘ</th>
            <th onclick="sortBy('status')">Status</th>
            <th id="th-asset" onclick="sortBy('assetStatus')" style="display:none">Asset</th>
          </tr></thead>
          <tbody id="tbody"></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- Deny Reasons tab -->
  <div class="tab-content" id="tab-reasons">
    <div class="section">
      <div class="section-title">Deny Reason Breakdown — click a reason to filter the client table</div>
      <div id="reasons-list"></div>
    </div>
  </div>

  <!-- Timeline tab -->
  <div class="tab-content" id="tab-timeline">
    <div class="section">
      <div class="section-title" id="timeline-title">7-Day Persistence Timeline</div>
      <div id="timeline"></div>
    </div>
  </div>

  <!-- Notifications tab -->
  <div class="tab-content" id="tab-notifications">
    <div class="section">
      <div class="section-title">Notification Center — Per-Site Email</div>
      <div style="font-size:12px;color:var(--text-muted);margin-bottom:10px">Select a site to compose a pre-populated remediation email:</div>
      <div class="site-selector" id="site-selector"></div>
      <div id="email-composer"></div>
    </div>
    <div class="section" style="margin-top:0">
      <div class="section-title">NOC Digest — All Sites</div>
      <div class="email-box">
        <div class="email-field"><label>Subject</label><input id="noc-subject" type="text"></div>
        <div class="email-field"><label>Body</label><textarea id="noc-body"></textarea></div>
        <div class="btn-row"><button class="btn btn-secondary" onclick="copy('noc-subject','noc-body')">Copy to Clipboard</button></div>
      </div>
    </div>
  </div>

  <!-- Help tab -->
  <div class="tab-content" id="tab-help">
    <div class="section">
      <div class="help-section">
        <h3>What is this report?</h3>
        <p>The Deny Log Intelligence report transforms raw RADIUS deny events from Mist Access Assurance into a client-centric view. Instead of seeing hundreds of rows for one broken device, you get <strong>one record per client</strong> — showing exactly why it's failing, how long it's been failing, and who else is hitting the same issue.</p>
        <p style="margin-top:8px">This file is self-contained. No internet connection required. You can email it directly to a colleague.</p>
      </div>
      <div class="help-section">
        <h3>Metric Cards</h3>
        <table class="help-table"><thead><tr><th>Card</th><th>What it means</th></tr></thead><tbody>
          <tr><td>Total Deny Events</td><td>Raw number of individual RADIUS deny events in the selected window. One broken client retrying 50× = 50 events.</td></tr>
          <tr><td>Unique Clients</td><td>Distinct devices (by MAC address) with at least one deny event. This is the number that matters for triage.</td></tr>
          <tr><td>Sites Affected</td><td>Distinct Mist sites with at least one failing client.</td></tr>
          <tr><td>Cert Failures</td><td>Clients whose primary failure is a certificate or TLS trust issue.</td></tr>
          <tr><td>Silent Failures</td><td>Clients that failed but have not retried in 8+ business hours. Invisible in real-time dashboards.</td></tr>
        </tbody></table>
      </div>
      <div class="help-section">
        <h3>Failure Categories</h3>
        <table class="help-table"><thead><tr><th>Category</th><th>Meaning</th><th>Common fix</th></tr></thead><tbody>
          <tr><td>🔐 Cert / TLS</td><td>Device doesn't trust the RADIUS server cert, or its own cert is missing/expired.</td><td>Import Mist cert via MDM. Organization → Access → Certificates.</td></tr>
          <tr><td>🔑 Credentials</td><td>Username/password rejected by IdP. Device itself may be fine.</td><td>Verify credentials, check account not locked, confirm IdP group grants access.</td></tr>
          <tr><td>📱 MAC / Policy</td><td>Hit the implicit deny — no NAC rule matched this device.</td><td>Confirm MAC is enrolled. Check NAC rules for this device type or SSID.</td></tr>
        </tbody></table>
      </div>
      <div class="help-section">
        <h3>Client Status</h3>
        <table class="help-table"><thead><tr><th>Status</th><th>Definition</th><th>Action</th></tr></thead><tbody>
          <tr><td><span class="badge badge-failing">failing</span></td><td>Deny event within the last 8 business hours. Actively retrying right now.</td><td>Immediate — user is likely on-site and unable to connect.</td></tr>
          <tr><td><span class="badge badge-silent">silent</span></td><td>Last deny was 8+ business hours ago (Mon–Fri 7am–7pm). No permit seen. User may have given up or gone home.</td><td>Proactive — reach out before they file a ticket Monday morning.</td></tr>
          <tr><td><span class="badge badge-resolved">resolved</span></td><td>A successful RADIUS permit was seen after the deny events.</td><td>No action needed. Monitor to confirm it doesn't reappear.</td></tr>
        </tbody></table>
      </div>
      <div class="help-section">
        <h3>Blast Radius</h3>
        <p>Measures how many clients share the <strong>exact same deny reason text</strong>. High blast = systemic issue (e.g. cert not deployed to all managed Windows devices).</p>
        <table class="help-table"><thead><tr><th>Score</th><th>Threshold</th><th>What it suggests</th></tr></thead><tbody>
          <tr><td><span class="badge badge-high">HIGH</span></td><td>5+ clients same reason</td><td>Systemic — policy change, failed MDM deployment, cert rollout missed a fleet.</td></tr>
          <tr><td><span class="badge badge-med">MED</span></td><td>3–4 clients</td><td>Small group — shared policy, same device model, or SSID config issue.</td></tr>
          <tr><td><span class="badge badge-low">LOW</span></td><td>1–2 clients</td><td>Isolated — specific to that device or user account.</td></tr>
        </tbody></table>
      </div>
      <div class="help-section">
        <h3>Frequently Asked Questions</h3>
        <table class="help-table"><tbody>
          <tr><td>Why is a client "silent" even though it failed today?</td><td>Silent needs 8+ <em>business</em> hours without a retry. An early-morning failure may not have crossed the threshold yet.</td></tr>
          <tr><td>Client shows "resolved" but user still can't connect.</td><td>A permit was seen in the window — but it may have failed again after. Try regenerating with a shorter lookback (3 days).</td></tr>
          <tr><td>Same MAC, different usernames.</td><td>Normal for shared or re-enrolled devices. MAC is the primary key; usernames shown for reference.</td></tr>
          <tr><td>Report shows 0 clients but I know there are failures.</td><td>Check that the token has org-level (not site-level) access. Site-level tokens miss cross-site events.</td></tr>
        </tbody></table>
      </div>
    </div>
  </div>

</div><!-- /.main -->

<script>
const REPORT = __REPORT_JSON__;
const LOOKBACK_DAYS   = REPORT.lookbackDays;
const BIZ_START       = 7, BIZ_END = 19, SILENCE_BIZ_HOURS = 8;
const CATEGORY_LABELS = { cert: 'Cert / TLS Issue', cred: 'Wrong Credentials', mac: 'MAC Auth / Policy Failure' };
const CATEGORY_REMEDIATION = {
  cert: 'Deploy updated client certificate via MDM. Cert chain must include the issuing CA trusted by the RADIUS server. Verify the client is configured to trust the Mist Authentication Service certificate (Organization > Access > Certificates).',
  cred: 'Verify username/password are correct and the account is not locked or expired. Check that the IdP group membership allows network access. Re-enroll the device if credentials were recently rotated.',
  mac:  'Confirm the device MAC address is enrolled in the correct MAC Auth list. Review NAC policy rules — the device is hitting the implicit deny. Check if the SSID or port policy requires a matching rule for this device type.',
};

let activeCat = null, activeReason = null, sortKey = 'blastCount', sortDir = -1;

// ── Dashboard render ─────────────────────────────────────────────────────────
function showDashboard() {
  const { orgName, generatedAt, lookbackDays, totalRawEvents, clients } = REPORT;
  document.getElementById('hdr-org').textContent  = orgName;
  document.getElementById('hdr-meta').textContent =
    `Generated ${fmtTs(generatedAt)} · ${lookbackDays}-day window · ${totalRawEvents.toLocaleString()} raw events`;
  document.getElementById('s-events-sub').textContent = `${lookbackDays}-day window`;

  const certC  = clients.filter(c => c.category === 'cert');
  const silent = clients.filter(c => c.status   === 'silent');
  const sites  = [...new Set(clients.map(c => c.site).filter(Boolean))];

  document.getElementById('s-events').textContent   = totalRawEvents.toLocaleString();
  document.getElementById('s-clients').textContent  = clients.length;
  document.getElementById('s-sites').textContent    = sites.length;
  document.getElementById('s-sites-names').textContent = sites.slice(0,3).join(', ') + (sites.length > 3 ? '…' : '');
  document.getElementById('s-cert').textContent     = certC.length;
  document.getElementById('s-cert-pct').textContent = clients.length ? `${Math.round(certC.length/clients.length*100)}% of clients` : '';
  document.getElementById('s-silent').textContent   = silent.length;
  document.getElementById('silent-help').title =
    `Silent = no retry in ${SILENCE_BIZ_HOURS} business hours (Mon–Fri ${BIZ_START}:00–${BIZ_END}:00). Weekend hours excluded.`;

  if (REPORT.hasAssets) {
    document.getElementById('card-managed').style.display = '';
    document.getElementById('s-managed').textContent      = REPORT.managedFailing;
    document.getElementById('s-managed-sub').textContent  =
      `${REPORT.unmanagedFailing} unmanaged also failing`;
    document.getElementById('f-asset').style.display  = '';
    document.getElementById('th-asset').style.display = '';
  }

  ['cert','cred','mac'].forEach(cat =>
    document.getElementById(`cc-${cat}`).textContent = clients.filter(c => c.category === cat).length
  );

  const siteEl = document.getElementById('f-site');
  siteEl.innerHTML = '<option value="">All Sites</option>';
  sites.sort().forEach(s => { const o = document.createElement('option'); o.value = s; o.textContent = s; siteEl.appendChild(o); });

  const reasonEl = document.getElementById('f-reason');
  reasonEl.innerHTML = '<option value="">All Deny Reasons</option>';
  REPORT.denyReasons.forEach(r => {
    const o = document.createElement('option');
    o.value = r.text;
    o.textContent = r.text.length > 70 ? r.text.slice(0,67) + '...' : r.text;
    reasonEl.appendChild(o);
  });

  document.getElementById('timeline-title').textContent = `${lookbackDays}-Day Persistence Timeline`;

  renderAlerts();
  renderTable();
  renderDenyReasons();
  renderTimeline();
  renderNotifications();
  renderNocDigest();
  switchTab('dashboard');
}

// ── Alerts ───────────────────────────────────────────────────────────────────
function renderAlerts() {
  const clients = REPORT.clients;
  const alerts  = [];
  const certC   = clients.filter(c => c.category === 'cert');
  if (certC.length) {
    const startDays = {};
    certC.forEach(c => {
      if (c.firstSeen) {
        const d = new Date(c.firstSeen * 1000).toLocaleDateString(undefined, { month: 'short', day: 'numeric' });
        startDays[d] = (startDays[d] || 0) + 1;
      }
    });
    const wave = Object.entries(startDays).filter(([,n]) => n >= 2).sort((a,b) => b[1]-a[1]);
    if (wave.length) {
      const [day, count] = wave[0];
      const waveSites = [...new Set(certC.map(c => c.site).filter(Boolean))];
      alerts.push({ type: 'cert', title: `Cert failure wave — ${count} clients started failing around ${day}`, body: `Affected sites: ${waveSites.join(', ')}.` });
    }
  }
  const silent = clients.filter(c => c.status === 'silent');
  if (silent.length) {
    alerts.push({ type: 'silent', title: `${silent.length} silent failure${silent.length > 1 ? 's' : ''} — clients that failed and stopped retrying`, body: `Last seen: ${silent.map(c => fmtDate(c.lastSeen)).join(', ')}.` });
  }
  const highBlast = clients.filter(c => c.blastScore === 'high');
  if (highBlast.length) {
    const groups = {};
    highBlast.forEach(c => { groups[c.primaryText] = (groups[c.primaryText] || []).concat(c); });
    Object.entries(groups).forEach(([reason, group]) => {
      const sitesAffected = [...new Set(group.map(c => c.site).filter(Boolean))];
      alerts.push({ type: 'high', title: `Systemic issue — ${group.length} clients sharing the same deny reason across ${sitesAffected.length} site${sitesAffected.length>1?'s':''}`, body: `Reason: "${reason.slice(0,120)}${reason.length > 120 ? '…' : ''}"` });
    });
  }
  document.getElementById('alerts').innerHTML = alerts.map(a => `
    <div class="alert-card ${a.type === 'silent' ? 'silent' : ''}">
      <div class="alert-title">${a.type === 'silent' ? '👻' : a.type === 'high' ? '🚨' : '⚠️'} ${a.title}</div>
      <div class="alert-body">${a.body}</div>
    </div>`).join('');
}

// ── Last-known location helper ────────────────────────────────────────────────
function locationLine(c) {
  const parts = [];
  if (c.ap)       parts.push(`AP: ${c.ap}`);
  else if (c.apMac) parts.push(`AP: ${c.apMac}`);
  if (c.portId)   parts.push(`Port: ${c.portId}`);
  if (c.switchMac && !c.portId) parts.push(`SW: ${c.switchMac}`);
  if (!parts.length) return '';
  return `<br><small style="color:var(--text-muted);font-size:10px" title="Last known location at time of deny">📍 ${parts.join(' · ')}</small>`;
}

// ── Asset cell helper ─────────────────────────────────────────────────────────
function assetCell(c) {
  const s = c.assetStatus || 'unknown';
  let html = `<span class="badge badge-${s}">${s}</span>`;
  if (s === 'managed') {
    const lines = [c.assetName, c.assetOwner, c.assetDept].filter(Boolean);
    if (lines.length) html += `<br><small style="color:var(--text-muted);font-size:10px">${lines.join(' · ')}</small>`;
  }
  return html;
}

// ── Client table ─────────────────────────────────────────────────────────────
function getFiltered() {
  let list   = [...REPORT.clients];
  const search  = document.getElementById('search').value.toLowerCase();
  const fCat    = document.getElementById('f-cat').value;
  const fSite   = document.getElementById('f-site').value;
  const fStatus = document.getElementById('f-status').value;
  const fReason = document.getElementById('f-reason').value;
  const fAsset  = document.getElementById('f-asset').value;

  if (activeCat)  list = list.filter(c => c.category === activeCat);
  if (activeReason) list = list.filter(c => c.allTexts.some(t => t.text === activeReason));
  if (fCat)    list = list.filter(c => c.category    === fCat);
  if (fSite)   list = list.filter(c => c.site        === fSite);
  if (fStatus) list = list.filter(c => c.status      === fStatus);
  if (fReason) list = list.filter(c => c.allTexts.some(t => t.text === fReason));
  if (fAsset)  list = list.filter(c => c.assetStatus === fAsset);
  if (search)  list = list.filter(c =>
    [c.mac, c.username, c.site, c.ssid, c.primaryText, c.diagnosis, c.assetName, c.assetOwner, c.assetDept, c.ap, c.portId, c.switchMac].some(v => (v||'').toLowerCase().includes(search))
  );

  const blastOrd = { high: 3, med: 2, low: 1 };
  list.sort((a, b) => {
    let av = a[sortKey], bv = b[sortKey];
    if (sortKey === 'blastScore') { av = blastOrd[av] || 0; bv = blastOrd[bv] || 0; }
    return (av < bv ? 1 : av > bv ? -1 : 0) * sortDir;
  });
  return list;
}

function sortBy(key) {
  sortDir = sortKey === key ? sortDir * -1 : -1;
  sortKey = key;
  renderTable();
}

function toggleCat(cat) {
  activeCat = activeCat === cat ? null : cat;
  ['cert','cred','mac'].forEach(c => document.getElementById(`cat-${c}`).classList.toggle('active-filter', c === activeCat));
  renderTable();
}

function clearFilters() {
  document.getElementById('search').value   = '';
  document.getElementById('f-cat').value    = '';
  document.getElementById('f-site').value   = '';
  document.getElementById('f-status').value = '';
  document.getElementById('f-reason').value = '';
  document.getElementById('f-asset').value  = '';
  activeCat = null; activeReason = null;
  ['cert','cred','mac'].forEach(c => document.getElementById(`cat-${c}`).classList.remove('active-filter'));
  document.querySelectorAll('.reason-row').forEach(r => r.classList.remove('active-reason'));
  renderTable();
}

function renderTable() {
  const list = getFiltered();
  document.getElementById('result-count').textContent = `${list.length} client${list.length !== 1 ? 's' : ''}`;
  const rows = list.map(c => {
    const label = (c.username && c.username !== c.mac)
      ? `${c.mac}<br><small style="color:var(--text-muted)">${c.username}</small>` : c.mac;
    const rawTip   = c.allTexts.map(t => `${t.text} (x${t.count})`).join('\n');
    const extra    = c.allTexts.length > 1 ? `<br><small style="color:var(--text-muted)">+${c.allTexts.length-1} more</small>` : '';
    const rawShort = (c.primaryText || '').length > 55 ? c.primaryText.slice(0,52) + '…' : c.primaryText;
    let reason;
    if (rawShort) {
      // Official RADIUS error — always shown as primary line
      reason = `<span title="${rawTip.replace(/"/g,'&quot;')}" style="cursor:help">${rawShort}</span>${extra}`;
      if (c.diagnosis) {
        // Suggested diagnosis + fix shown below as a secondary line
        const diagShort = c.diagnosis.length > 60 ? c.diagnosis.slice(0,57) + '…' : c.diagnosis;
        const fixTip    = (c.specificFix || '').replace(/"/g,'&quot;');
        reason += `<br><small title="${fixTip}" style="color:var(--accent);cursor:help">💡 ${diagShort}</small>`;
      }
    } else {
      reason = '—';
    }
    return `<tr>
      <td class="mac-cell">${label||'—'}</td>
      <td>${c.site||'—'}${c.ssid?`<br><small style="color:var(--text-muted)">${c.ssid}</small>`:''}${locationLine(c)}</td>
      <td><span class="badge badge-${c.category}">${c.categoryLabel}</span></td>
      <td style="font-size:12px;line-height:1.4">${reason}</td>
      <td>${fmtDate(c.firstSeen)}</td>
      <td>${fmtDate(c.lastSeen)}</td>
      <td style="text-align:center;font-weight:700">${c.daysFailing}</td>
      <td style="text-align:center">${c.attempts}</td>
      <td title="${c.blastCount} client${c.blastCount!==1?'s':''} share this deny reason">
        <span class="badge badge-${c.blastScore}">${c.blastScore.toUpperCase()}</span>
        <br><small style="color:var(--text-muted)">${c.blastCount} client${c.blastCount!==1?'s':''}</small>
      </td>
      <td><span class="badge badge-${c.status}">${c.status}</span></td>
      ${REPORT.hasAssets ? `<td>${assetCell(c)}</td>` : ''}
    </tr>`;
  }).join('');
  document.getElementById('tbody').innerHTML = rows ||
    `<tr><td colspan="${REPORT.hasAssets ? 11 : 10}" style="text-align:center;color:var(--text-muted);padding:32px">No clients match the current filters.</td></tr>`;
}

// ── Deny reasons ─────────────────────────────────────────────────────────────
function renderDenyReasons() {
  const reasons = REPORT.denyReasons;
  if (!reasons.length) { document.getElementById('reasons-list').innerHTML = '<p style="color:var(--text-muted)">No deny events found.</p>'; return; }
  const max = reasons[0].clientCount;
  document.getElementById('reasons-list').innerHTML = reasons.map((r, idx) => {
    const barW = Math.max(4, Math.round(r.clientCount / max * 120));
    const cat  = r.clients[0]?.category || 'mac';
    const macs = r.clients.map(c => c.username && c.username !== c.mac ? c.username : c.mac).join(', ');
    const col  = cat === 'cert' ? 'var(--cert)' : cat === 'cred' ? 'var(--cred)' : 'var(--mac)';
    const safeText = r.text.replace(/\\/g,'\\\\').replace(/'/g,"\\'").replace(/"/g,'&quot;');
    return `<div class="reason-row" id="rr-${idx}" onclick="filterByReason('${safeText}',${idx})">
      <div class="reason-bar-wrap">
        <div class="reason-bar" style="width:${barW}px;background:${col}"></div>
        <span class="reason-client-count">${r.clientCount} client${r.clientCount!==1?'s':''} · ${r.totalEvents} events</span>
      </div>
      <div class="reason-text">${r.text}<div class="reason-macs">${macs}</div></div>
      <span class="badge badge-${cat}">${cat}</span>
    </div>`;
  }).join('');
}

function filterByReason(text, idx) {
  activeReason = activeReason === text ? null : text;
  document.querySelectorAll('.reason-row').forEach(r => r.classList.remove('active-reason'));
  if (activeReason) document.getElementById(`rr-${idx}`)?.classList.add('active-reason');
  switchTab('dashboard');
  renderTable();
}

// ── Timeline ─────────────────────────────────────────────────────────────────
function renderTimeline() {
  const { clients, dayLabels } = REPORT;
  const hdr = dayLabels.map(d => {
    const dt = new Date(d + 'T12:00:00Z');
    return `<div class="timeline-day-label">${dt.toLocaleDateString(undefined,{month:'short',day:'numeric'})}</div>`;
  }).join('');
  const rows = clients.map(c => {
    const lbl   = c.username && c.username !== c.mac ? `${c.mac} / ${c.username}` : c.mac;
    const cells = dayLabels.map(d => {
      const n = c.activity[d] || 0;
      return `<div class="timeline-cell ${n > 0 ? c.category : ''}" title="${n > 0 ? `${lbl} · ${d} · ${n} attempts` : ''}"></div>`;
    }).join('');
    return `<div class="timeline-row">
      <div class="timeline-client-label" title="${lbl}">${c.username || c.mac}</div>
      <div class="timeline-cells">${cells}</div>
    </div>`;
  }).join('');
  document.getElementById('timeline').innerHTML = `<div class="timeline-header">${hdr}</div>${rows}`;
}

// ── Notifications ─────────────────────────────────────────────────────────────
let activeSite = null;
function renderNotifications() {
  const sites = [...new Set(REPORT.clients.map(c => c.site).filter(Boolean))].sort();
  document.getElementById('site-selector').innerHTML = sites.map(s =>
    `<div class="site-btn" id="sb-${s.replace(/\W/g,'_')}" onclick="selectSite('${s.replace(/'/g,"\\'")}')"> ${s}</div>`
  ).join('');
  if (sites.length) selectSite(sites[0]);
}

function selectSite(site) {
  activeSite = site;
  document.querySelectorAll('.site-btn').forEach(b => b.classList.remove('active'));
  document.getElementById(`sb-${site.replace(/\W/g,'_')}`)?.classList.add('active');
  const sc   = REPORT.clients.filter(c => c.site === site);
  const byCat = { cert: [], cred: [], mac: [] };
  sc.forEach(c => (byCat[c.category] || byCat.mac).push(c));
  const subject = `[Mist NAC] Access Failures — ${site} — ${new Date().toLocaleDateString()}`;
  let body = `Hi team,\n\nWe are seeing RADIUS authentication failures at ${site}.\nThis report was generated from the Mist Deny Log Intelligence tool.\n\n`;
  body += `SUMMARY\n${'─'.repeat(50)}\nTotal affected clients: ${sc.length}\nWindow: Last ${REPORT.lookbackDays} days\n\n`;
  ['cert','cred','mac'].forEach(cat => {
    const cc = byCat[cat];
    if (!cc.length) return;
    body += `${CATEGORY_LABELS[cat].toUpperCase()} (${cc.length} client${cc.length>1?'s':''})\n${'─'.repeat(40)}\n`;
    cc.forEach(c => {
      body += `  • MAC: ${c.mac}`;
      if (c.username && c.username !== c.mac) body += `  User: ${c.username}`;
      body += `  — ${c.daysFailing} day${c.daysFailing>1?'s':''}, ${c.attempts} attempts (${c.status})\n`;
      if (c.diagnosis)    body += `    Cause: ${c.diagnosis}\n`;
      else if (c.primaryText) body += `    Error: ${c.primaryText}\n`;
      if (c.specificFix)  body += `    Fix:   ${c.specificFix}\n`;
      const loc = [];
      if (c.ap || c.apMac)  loc.push(`AP: ${c.ap || c.apMac}`);
      if (c.portId)         loc.push(`Port: ${c.portId}`);
      if (c.switchMac && !c.portId) loc.push(`Switch: ${c.switchMac}`);
      if (loc.length)       body += `    Last seen: ${loc.join(' · ')}\n`;
    });
    body += `\nRemediation: ${CATEGORY_REMEDIATION[cat]}\n\n`;
  });
  body += `Reply to this email or contact the NOC for assistance.\n\nRegards,\nNetwork Operations Center`;
  document.getElementById('email-composer').innerHTML = `
    <div class="email-box">
      <div class="email-field"><label>To</label><input type="text" placeholder="building-tech@site.edu" id="e-to"></div>
      <div class="email-field"><label>Subject</label><input type="text" id="e-subject" value="${subject.replace(/"/g,'&quot;')}"></div>
      <div class="email-field"><label>Body (editable)</label><textarea id="e-body">${body}</textarea></div>
      <div class="btn-row">
        <button class="btn btn-primary" onclick="copy('e-subject','e-body')">Copy to Clipboard</button>
        <button class="btn btn-secondary" onclick="openMailto()">Open in Mail App</button>
      </div>
    </div>`;
}

function openMailto() {
  const s = encodeURIComponent(document.getElementById('e-subject').value);
  const b = encodeURIComponent(document.getElementById('e-body').value);
  window.location.href = `mailto:?subject=${s}&body=${b}`;
}

// ── NOC Digest ─────────────────────────────────────────────────────────────
function renderNocDigest() {
  const { clients, lookbackDays, totalRawEvents, orgName } = REPORT;
  const sites   = [...new Set(clients.map(c => c.site).filter(Boolean))].sort();
  const subject = `[NOC Digest] Mist NAC Deny Report — ${new Date().toLocaleDateString()} — ${sites.length} site${sites.length>1?'s':''}`;
  let body = `NOC DIGEST — MIST ACCESS ASSURANCE DENY REPORT\nOrg: ${orgName}\nGenerated: ${new Date().toLocaleString()}\nWindow: Last ${lookbackDays} days\n\n`;
  body += `ORGANIZATION SUMMARY\n${'='.repeat(50)}\n`;
  body += `Total deny events:   ${totalRawEvents.toLocaleString()}\nUnique clients:      ${clients.length}\nSites affected:      ${sites.length}\n`;
  body += `Cert failures:       ${clients.filter(c=>c.category==='cert').length}\nCred failures:       ${clients.filter(c=>c.category==='cred').length}\nMAC/policy failures: ${clients.filter(c=>c.category==='mac').length}\nSilent failures:     ${clients.filter(c=>c.status==='silent').length}\n\n`;
  sites.forEach(site => {
    const sc   = clients.filter(c => c.site === site);
    const high = sc.filter(c => c.blastScore === 'high');
    body += `SITE: ${site}\n${'─'.repeat(40)}\n`;
    body += `  Clients: ${sc.length} | Cert: ${sc.filter(c=>c.category==='cert').length} | Cred: ${sc.filter(c=>c.category==='cred').length} | MAC: ${sc.filter(c=>c.category==='mac').length}\n`;
    if (high.length) body += `  HIGH blast radius: ${high.map(c=>c.mac).join(', ')}\n`;
    body += `  Action: ${sc.length > 0 ? 'Notify building tech. See per-site email.' : 'None required.'}\n\n`;
  });
  body += `Full details in the attached HTML report.\n`;
  document.getElementById('noc-subject').value = subject;
  document.getElementById('noc-body').value    = body;
}

// ── Tabs + utils ──────────────────────────────────────────────────────────────
function switchTab(name) {
  const names = ['dashboard','reasons','timeline','notifications','help'];
  document.querySelectorAll('.tab').forEach((t,i) => t.classList.toggle('active', names[i] === name));
  document.querySelectorAll('.tab-content').forEach(t => t.classList.toggle('active', t.id === `tab-${name}`));
}

function copy(subjectId, bodyId) {
  const s = document.getElementById(subjectId)?.value || '';
  const b = document.getElementById(bodyId)?.value   || '';
  navigator.clipboard.writeText(`Subject: ${s}\n\n${b}`)
    .then(() => alert('Copied to clipboard!'))
    .catch(() => alert('Copy failed — please select and copy manually.'));
}

function fmtTs(ts) {
  if (!ts) return '—';
  return new Date(ts * 1000).toLocaleString(undefined, { month: 'short', day: 'numeric', hour: '2-digit', minute: '2-digit' });
}
function fmtDate(ts) {
  if (!ts) return '—';
  return new Date(ts * 1000).toLocaleDateString(undefined, { month: 'short', day: 'numeric' });
}

window.addEventListener('DOMContentLoaded', () => {
  showDashboard();
  // Drag-and-drop CSV onto the whole page
  document.addEventListener('dragover', e => { e.preventDefault(); document.body.classList.add('drop-zone-active'); });
  document.addEventListener('dragleave', e => { if (!e.relatedTarget) document.body.classList.remove('drop-zone-active'); });
  document.addEventListener('drop', e => {
    e.preventDefault();
    document.body.classList.remove('drop-zone-active');
    const file = [...e.dataTransfer.files].find(f => f.name.match(/\.csv$/i));
    if (file) handleCSVFile(file);
    else showToast('Drop a .csv file to import an asset list.', 'warn');
  });
});

// ── Toast ─────────────────────────────────────────────────────────────────────
let toastTimer;
function showToast(msg, type = 'ok') {
  const el = document.getElementById('toast');
  el.textContent = msg;
  el.className   = `toast toast-${type} show`;
  clearTimeout(toastTimer);
  toastTimer = setTimeout(() => el.classList.remove('show'), 4000);
}

// ── Asset CSV import ──────────────────────────────────────────────────────────
const _MAC_KW  = ['mac','address','hardware','hwaddr','ethernet','wifi','wireless','wlan','bssid','physical'];
const _NAME_KW = ['name','hostname','computer','device','asset','label','computername'];
const _OWN_KW  = ['user','owner','assigned','person','email','upn','login'];
const _DEPT_KW = ['dept','department','group','division','team','ou','org','unit'];

function normalizeMac(s) {
  return (s || '').replace(/[:\-.\s]/g, '').toLowerCase();
}

function bestColIdx(headers, keywords) {
  const hl = headers.map(h => h.toLowerCase().replace(/[\s_\-]/g, ''));
  for (const kw of keywords) {
    const i = hl.findIndex(h => h.includes(kw));
    if (i >= 0) return i;
  }
  return -1;
}

function parseCSV(text) {
  if (text.charCodeAt(0) === 0xFEFF) text = text.slice(1);  // strip BOM
  const lines = text.split(/\r?\n/).filter(l => l.trim());
  if (lines.length < 2) return null;
  function parseLine(line) {
    const fields = []; let cur = '', inQ = false;
    for (const ch of line) {
      if (ch === '"') { inQ = !inQ; continue; }
      if (ch === ',' && !inQ) { fields.push(cur.trim()); cur = ''; continue; }
      cur += ch;
    }
    fields.push(cur.trim());
    return fields;
  }
  const headers = parseLine(lines[0]);
  const rows    = lines.slice(1).map(parseLine);
  return { headers, rows };
}

function handleCSVFile(file) {
  if (!file) return;
  // Reset input so same file can be re-selected
  document.getElementById('csv-file-input').value = '';
  const reader = new FileReader();
  reader.onload = e => processCSVText(e.target.result, file.name);
  reader.readAsText(file);
}

function processCSVText(text, filename) {
  const parsed = parseCSV(text);
  if (!parsed) { showToast('Could not parse CSV — file appears empty.', 'err'); return; }

  const macIdx  = bestColIdx(parsed.headers, _MAC_KW);
  if (macIdx < 0) {
    showToast(`No MAC address column found.\nColumns: ${parsed.headers.join(', ')}`, 'err');
    return;
  }

  const nameIdx = bestColIdx(parsed.headers, _NAME_KW);
  const ownIdx  = bestColIdx(parsed.headers, _OWN_KW);
  const deptIdx = bestColIdx(parsed.headers, _DEPT_KW);

  const assetMap = {};
  let loaded = 0;
  for (const row of parsed.rows) {
    const norm = normalizeMac(row[macIdx] || '');
    if (!norm || norm.length !== 12) continue;
    assetMap[norm] = {
      name:  nameIdx >= 0 ? (row[nameIdx]  || '') : '',
      owner: ownIdx  >= 0 ? (row[ownIdx]   || '') : '',
      dept:  deptIdx >= 0 ? (row[deptIdx]  || '') : '',
    };
    loaded++;
  }

  if (!loaded) { showToast('No valid MAC addresses found in the CSV.', 'err'); return; }
  applyAssets(assetMap, filename, loaded);
}

function applyAssets(assetMap, filename, loadedCount) {
  let managed = 0, unmanaged = 0;
  REPORT.clients.forEach(c => {
    const info = assetMap[normalizeMac(c.mac || '')];
    if (info) {
      c.assetStatus = 'managed';
      c.assetName   = info.name;
      c.assetOwner  = info.owner;
      c.assetDept   = info.dept;
      managed++;
    } else {
      c.assetStatus = 'unmanaged';
      c.assetName = c.assetOwner = c.assetDept = '';
      unmanaged++;
    }
  });

  REPORT.hasAssets      = true;
  REPORT.managedFailing = managed;
  REPORT.unmanagedFailing = unmanaged;

  // Show asset bar
  document.getElementById('asset-bar').style.display = '';
  document.getElementById('asset-bar-text').textContent =
    `Asset list: ${filename} · ${loadedCount} devices loaded · ${managed} matched (managed) · ${unmanaged} unmatched (unmanaged)`;

  // Show metric card + filter + column
  document.getElementById('card-managed').style.display = '';
  document.getElementById('s-managed').textContent      = managed;
  document.getElementById('s-managed-sub').textContent  = `${unmanaged} unmanaged also failing`;
  document.getElementById('f-asset').style.display      = '';
  document.getElementById('th-asset').style.display     = '';

  renderTable();
  showToast(`✓ Asset list loaded — ${managed} managed devices matched`, 'ok');
}

function clearAssets() {
  REPORT.clients.forEach(c => {
    c.assetStatus = 'unknown';
    c.assetName = c.assetOwner = c.assetDept = '';
  });
  REPORT.hasAssets = false;
  REPORT.managedFailing = 0;
  REPORT.unmanagedFailing = 0;

  document.getElementById('asset-bar').style.display  = 'none';
  document.getElementById('card-managed').style.display = 'none';
  document.getElementById('f-asset').style.display    = 'none';
  document.getElementById('f-asset').value            = '';
  document.getElementById('th-asset').style.display   = 'none';
  renderTable();
  showToast('Asset list removed.', 'warn');
}
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def open_file(path):
    abs_path = os.path.abspath(path)
    try:
        if sys.platform == "darwin":
            subprocess.run(["open", abs_path], check=False)
        elif sys.platform == "win32":
            os.startfile(abs_path)
        else:
            subprocess.run(["xdg-open", abs_path], check=False)
    except Exception:
        pass


def main():
    # ── Dependency check ──────────────────────────────────────────────────────
    if requests is None:
        print("\nERROR: 'requests' library is not installed.")
        print("Run:  pip install requests openpyxl")
        sys.exit(1)

    print("\n" + "═" * 55)
    print("  Mist Access Assurance — Deny Log Intelligence")
    print("═" * 55)

    # ── Token ─────────────────────────────────────────────────────────────────
    token = getpass.getpass("\nMist API Token (input hidden): ").strip()
    if not token:
        print("Token is required. Exiting.")
        sys.exit(1)

    # ── Cloud region ──────────────────────────────────────────────────────────
    print("\nCloud Region:")
    for i, (name, _) in enumerate(REGIONS, 1):
        print(f"  {i}. {name}")
    choice = input("Select region [1]: ").strip() or "1"
    try:
        base_url = REGIONS[int(choice) - 1][1]
    except (ValueError, IndexError):
        base_url = REGIONS[0][1]
    print(f"  → {base_url}")

    # ── Lookback window ───────────────────────────────────────────────────────
    days_input = input("\nLookback window in days [7]: ").strip() or "7"
    try:
        lookback_days = int(days_input)
    except ValueError:
        lookback_days = 7
    if lookback_days < 1 or lookback_days > 30:
        print("Clamping lookback to 7 days.")
        lookback_days = 7

    # ── Authenticate ──────────────────────────────────────────────────────────
    print("\n" + "─" * 40)
    print("Authenticating...")
    try:
        org_info = fetch_org_info(token, base_url)
    except requests.exceptions.HTTPError as e:
        print(f"\n  ERROR: API returned {e.response.status_code}.")
        if e.response.status_code == 401:
            print("  Token is invalid or expired. Check your API token.")
        sys.exit(1)
    except Exception as e:
        print(f"\n  ERROR: {e}")
        sys.exit(1)

    org_id   = org_info["org_id"]
    org_name = org_info["org_name"]
    print(f"  ✓  Org: {org_name}  ({org_id})")

    # ── Fetch sites ───────────────────────────────────────────────────────────
    print("Fetching sites...")
    site_map = fetch_sites(token, org_id, base_url)
    print(f"  ✓  {len(site_map)} sites found")

    # ── Fetch events ──────────────────────────────────────────────────────────
    print(f"Fetching deny events ({lookback_days}-day window)...")
    events = fetch_events(token, org_id, lookback_days, base_url)

    # ── Aggregate ─────────────────────────────────────────────────────────────
    print("Aggregating client records...")
    clients, day_labels, deny_reasons = aggregate_events(events, site_map, lookback_days)

    deny_event_count = sum(1 for e in events if e.get("type") in DENY_EVENT_TYPES)

    print(f"  ✓  {len(clients)} unique clients · {deny_event_count:,} deny events")

    # ── Build report object ────────────────────────────────────────────────────
    report = {
        "orgName":        org_name,
        "orgId":          org_id,
        "generatedAt":    datetime.now(tz=timezone.utc).timestamp(),
        "lookbackDays":   lookback_days,
        "totalRawEvents": deny_event_count,
        "clients":        clients,
        "denyReasons":    deny_reasons,
        "dayLabels":      day_labels,
        "hasAssets":      False,
        "managedFailing": 0,
        "unmanagedFailing": 0,
    }

    # ── Output paths ──────────────────────────────────────────────────────────
    ts         = datetime.now().strftime("%Y%m%d_%H%M")
    html_path  = f"deny_report_{ts}.html"
    xlsx_path  = f"deny_report_{ts}.xlsx"

    # ── HTML ──────────────────────────────────────────────────────────────────
    print(f"\nWriting HTML  → {html_path}")
    html = HTML_TEMPLATE.replace("__REPORT_JSON__", json.dumps(report))
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)

    # ── Excel ─────────────────────────────────────────────────────────────────
    if openpyxl:
        print(f"Writing Excel → {xlsx_path}")
        build_excel(report, xlsx_path)
    else:
        print("  ⚠  openpyxl not installed — skipping Excel.")
        print("     Run:  pip install openpyxl")
        xlsx_path = None

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "═" * 55)
    print("  ✓  Done!")
    print(f"     Org:          {org_name}")
    print(f"     Events:       {deny_event_count:,}")
    print(f"     Clients:      {len(clients)}")
    print(f"     Sites:        {len({c['site'] for c in clients if c.get('site')})}")
    print(f"     HTML:         {html_path}")
    if xlsx_path:
        print(f"     Excel:        {xlsx_path}")
    print("═" * 55 + "\n")

    # ── Open both files ───────────────────────────────────────────────────────
    open_file(html_path)
    if xlsx_path:
        open_file(xlsx_path)


if __name__ == "__main__":
    main()
